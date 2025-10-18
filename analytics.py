import os
import psycopg2
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from dotenv import load_dotenv

load_dotenv()
os.makedirs("charts", exist_ok=True)
os.makedirs("exports", exist_ok=True)

plt.style.use('seaborn-v0_8')
COLORS = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', '#DDA0DD']

def run_query(query):
    conn = psycopg2.connect(
        dbname=os.getenv("DB_NAME"), user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"), host=os.getenv("DB_HOST"), port=os.getenv("DB_PORT")
    )
    df = pd.read_sql_query(query, conn)
    conn.close()
    return df

def create_pie_chart():
    query = """
    SELECT c.customer_state, COUNT(*) AS total_orders
    FROM olist_orders o
    JOIN olist_customers c ON o.customer_id = c.customer_id
    JOIN olist_order_payments p ON o.order_id = p.order_id
    GROUP BY c.customer_state
    ORDER BY total_orders DESC
    LIMIT 8;
    """
    df = run_query(query)
    plt.figure(figsize=(10, 8))
    plt.pie(df['total_orders'], labels=df['customer_state'], autopct='%1.1f%%', colors=COLORS)
    plt.title('Distribution of Orders by Customer State')
    plt.savefig('charts/pie_orders_by_state.png')
    plt.close()
    print(f"Pie Chart: {len(df)} rows")

def create_bar_chart():
    query = """
    SELECT p.payment_type, COUNT(*) as total_orders, ROUND(AVG(p.payment_value), 2) as avg_payment
    FROM olist_orders o
    JOIN olist_customers c ON o.customer_id = c.customer_id
    JOIN olist_order_payments p ON o.order_id = p.order_id
    GROUP BY p.payment_type
    ORDER BY total_orders DESC;
    """
    df = run_query(query)
    plt.figure(figsize=(12, 6))
    plt.bar(df['payment_type'], df['total_orders'], color=COLORS[0])
    plt.title('Total Orders by Payment Type')
    plt.xlabel('Payment Type')
    plt.ylabel('Number of Orders')
    plt.savefig('charts/bar_orders_by_payment.png')
    plt.close()
    print(f"Bar Chart: {len(df)} rows")

def create_horizontal_bar_chart():
    query = """
    SELECT c.customer_state, COUNT(DISTINCT c.customer_id) as total_customers
    FROM olist_customers c
    JOIN olist_orders o ON c.customer_id = o.customer_id
    JOIN olist_order_payments p ON o.order_id = p.order_id
    GROUP BY c.customer_state
    ORDER BY total_customers DESC
    LIMIT 10;
    """
    df = run_query(query)
    plt.figure(figsize=(12, 6))
    plt.barh(df['customer_state'], df['total_customers'], color=COLORS[1])
    plt.title('Top 10 States by Number of Customers')
    plt.xlabel('Number of Customers')
    plt.savefig('charts/barh_customers_by_state.png')
    plt.close()
    print(f"Horizontal Bar Chart: {len(df)} rows")

def create_line_chart():
    query = """
    SELECT DATE_TRUNC('month', o.order_purchase_timestamp) as month,
           COUNT(DISTINCT o.order_id) as monthly_orders,
           ROUND(AVG(p.payment_value), 2) as avg_payment
    FROM olist_orders o
    JOIN olist_customers c ON o.customer_id = c.customer_id
    JOIN olist_order_payments p ON o.order_id = p.order_id
    WHERE o.order_purchase_timestamp IS NOT NULL
    GROUP BY month
    ORDER BY month;
    """
    df = run_query(query)
    df['month'] = pd.to_datetime(df['month'])
    plt.figure(figsize=(14, 6))
    plt.plot(df['month'], df['monthly_orders'], marker='o', color=COLORS[2])
    plt.title('Monthly Order Trends')
    plt.xlabel('Month')
    plt.ylabel('Number of Orders')
    plt.xticks(rotation=45)
    plt.savefig('charts/line_monthly_trends.png')
    plt.close()
    print(f"Line Chart: {len(df)} rows")

def create_histogram():
    query = """
    SELECT oi.price
    FROM olist_order_items oi
    JOIN olist_orders o ON oi.order_id = o.order_id
    JOIN olist_products p ON oi.product_id = p.product_id
    WHERE oi.price < 500 AND oi.price > 0;
    """
    df = run_query(query)
    plt.figure(figsize=(12, 6))
    plt.hist(df['price'], bins=30, color=COLORS[3], alpha=0.7)
    plt.title('Distribution of Product Prices')
    plt.xlabel('Product Price (R$)')
    plt.ylabel('Frequency')
    plt.savefig('charts/hist_product_prices.png')
    plt.close()
    print(f"Histogram: {len(df)} rows")

def create_scatter_plot():
    query = """
    SELECT oi.price, oi.freight_value, p.payment_value
    FROM olist_order_items oi
    JOIN olist_orders o ON oi.order_id = o.order_id
    JOIN olist_order_payments p ON o.order_id = p.order_id
    WHERE oi.price < 200 AND oi.freight_value < 50
    AND oi.price > 0 AND oi.freight_value > 0
    LIMIT 1000;
    """
    df = run_query(query)
    plt.figure(figsize=(10, 6))
    plt.scatter(df['price'], df['freight_value'], alpha=0.6, color=COLORS[4])
    plt.title('Price vs Freight Value')
    plt.xlabel('Product Price (R$)')
    plt.ylabel('Freight Value (R$)')
    plt.savefig('charts/scatter_price_vs_freight.png')
    plt.close()
    print(f"Scatter Plot: {len(df)} rows")

def create_all_visualizations():
    create_pie_chart()
    create_bar_chart()
    create_horizontal_bar_chart()
    create_line_chart()
    create_histogram()
    create_scatter_plot()

def create_time_slider_chart():
    query = """
    SELECT DATE_TRUNC('month', o.order_purchase_timestamp) as month,
           c.customer_state, COUNT(*) as order_count,
           ROUND(AVG(p.payment_value), 2) as avg_payment
    FROM olist_orders o
    JOIN olist_customers c ON o.customer_id = c.customer_id
    JOIN olist_order_payments p ON o.order_id = p.order_id
    WHERE o.order_purchase_timestamp IS NOT NULL
    AND c.customer_state IS NOT NULL
    GROUP BY month, c.customer_state
    HAVING COUNT(*) > 1
    ORDER BY month;
    """
    df = run_query(query)
    
    if df.empty:
        print("No data for time slider")
        return
        
    df['month'] = pd.to_datetime(df['month']).dt.strftime('%Y-%m')
    
    print(f"Time Slider Data: {len(df)} rows, {df['month'].nunique()} unique months")
    print(f"Months available: {sorted(df['month'].unique())}")
    
    fig = px.scatter(df, x="customer_state", y="order_count", size="avg_payment",
                    color="customer_state", animation_frame="month", 
                    title="Orders Over Time by State",
                    labels={"customer_state": "State", "order_count": "Orders", "avg_payment": "Avg Payment"})
    fig.show()
    
def export_to_excel():
    queries = {
        "Order_Summary": """
            SELECT c.customer_state, COUNT(*) as total_orders,
                   ROUND(AVG(p.payment_value), 2) as avg_payment
            FROM olist_orders o
            JOIN olist_customers c ON o.customer_id = c.customer_id
            JOIN olist_order_payments p ON o.order_id = p.order_id
            GROUP BY c.customer_state
            ORDER BY total_orders DESC;
        """,
        "Payment_Analysis": """
            SELECT p.payment_type, COUNT(*) as transaction_count,
                   ROUND(AVG(p.payment_value), 2) as avg_payment,
                   MIN(p.payment_value) as min_payment,
                   MAX(p.payment_value) as max_payment
            FROM olist_order_payments p
            JOIN olist_orders o ON p.order_id = o.order_id
            JOIN olist_customers c ON o.customer_id = c.customer_id
            GROUP BY p.payment_type
            ORDER BY transaction_count DESC;
        """
    }
    
    filename = "exports/olist_report.xlsx"
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        for sheet_name, query in queries.items():
            df = run_query(query)
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    
    wb = load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        for col in ws.iter_cols(min_row=2):
            if all(isinstance(cell.value, (int, float)) for cell in col if cell.value):
                col_letter = col[0].column_letter
                rule = ColorScaleRule(start_type="min", start_color="FFAA0000", mid_type="percentile", mid_value=50, mid_color="FFFFFF00", end_type="max", end_color="FF00AA00")
                ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)
    wb.save(filename)

if __name__ == "__main__":
    create_all_visualizations()
    create_time_slider_chart()
    export_to_excel()